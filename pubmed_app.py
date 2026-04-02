import streamlit as st
from Bio import Entrez
import pandas as pd
import re
from time import sleep
from io import BytesIO
import logging
import feedparser
import urllib.parse
from zipfile import ZipFile

# ---------- CONFIG ----------
Entrez.email = "your_email@example.com"  # Replace with your actual email
logging.basicConfig(level=logging.INFO)

# ---------- CONSTANTS ----------
PERSONAL_EMAIL_DOMAINS = ['gmail.com', 'yahoo.com', 'hotmail.com', 'outlook.com', 'aol.com']
BATCH_SIZE = 100

# ---------- FUNCTIONS ----------
def extract_email(text):
    match = re.search(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", text)
    return match.group(0).rstrip('.') if match else ""

def extract_university_name(affiliation_text):
    keywords = ['university', 'institute', 'college', 'school', 'center', 'centre', 'hospital']
    parts = [p.strip() for p in re.split(r'[;,]', affiliation_text.lower()) if p.strip()]
    matches = [p.title() for p in parts if any(k in p for k in keywords)]
    return matches[-1] if matches else ""

def format_mla(authors, title, journal, volume, issue, year, pages, doi):
    author_str = ", ".join(authors)
    return f"{author_str}. \"{title}.\" *{journal}*, vol. {volume}, no. {issue}, {year}, pp. {pages}. doi:{doi}"

def extract_doi(elocations):
    if isinstance(elocations, dict):
        elocations = [elocations]
    elif not isinstance(elocations, list):
        elocations = [elocations]
    for eloc in elocations:
        try:
            if hasattr(eloc, 'attributes') and eloc.attributes.get("EIdType") == "doi":
                return str(eloc)
        except Exception:
            continue
    return "N/A"

def get_google_news(query, max_articles=5):
    encoded_query = urllib.parse.quote(query)
    url = f'https://news.google.com/rss/search?q={encoded_query}'
    feed = feedparser.parse(url)
    news = []

    if not feed.entries:
        return []

    for entry in feed.entries[:max_articles]:
        news.append({
            'title': entry.title,
            'link': entry.link,
            'published': entry.published
        })
    return news

# ---------- STREAMLIT UI ----------
st.set_page_config(page_title="IOTA Tools", layout="wide")

menu = st.sidebar.selectbox("🔍 Select Tool", [
    "PubMed Article Extractor",
    "Google News Search",
    "Excel Splitter",
    "Yahoo Finance Company Lookup",
    "Excel Merger-Flatten Viewer",
    "Fuzzy Name Matcher",
    "Categories Lister",
    "Patent Scraper"   # ✅ NEW TOOL
])
# ==========================
# 🚀 PubMed Article Extractor
# ==========================
if menu == "PubMed Article Extractor":
    st.title("🔬 Auto Author/Article Extractor")

    search_term = st.text_input(
        "Enter your PubMed search term",
        '(Human Biology) AND ("united states"[Affiliation] OR USA[Affiliation]) AND (2022[Date - Publication])'
    )

    selected_countries = st.multiselect(
        "🌍 Select Countries (match in affiliation text)",
        options=[
            "USA", "United States", "United Kingdom", "Germany", "India", "Canada", "Australia",
            "France", "China", "Japan", "Brazil", "Italy", "Spain", "Netherlands", "Switzerland"
        ],
        default=["USA", "United States"]
    )

    retstart = st.number_input("Start from record number", min_value=0, value=0, step=100)
    retmax = st.number_input("How many records to fetch", min_value=10, max_value=10000, value=100, step=10)

    start_button = st.button("Fetch Articles")

    if not selected_countries:
        st.warning("⚠️ Please select at least one country to proceed.")

    if start_button and selected_countries:
        st.info(f"🔎 Searching PubMed (records {retstart} to {retstart + retmax - 1})...")
        try:
            search_handle = Entrez.esearch(
                db="pubmed",
                term=search_term,
                retstart=retstart,
                retmax=retmax
            )
            search_results = Entrez.read(search_handle)
        except Exception as e:
            st.error(f"❌ Failed to search PubMed: {e}")
            st.stop()

        pmids = search_results.get("IdList", [])
        if not pmids:
            st.error("No articles found for the given query and range.")
            st.stop()

        data = []
        progress = st.progress(0)

        for i, start in enumerate(range(0, len(pmids), BATCH_SIZE)):
            end = min(start + BATCH_SIZE, len(pmids))
            batch_pmids = pmids[start:end]

            try:
                fetch_handle = Entrez.efetch(db="pubmed", id=batch_pmids, rettype="xml")
                articles = Entrez.read(fetch_handle)
            except Exception as e:
                st.error(f"❌ Failed to fetch batch {start}-{end}: {e}")
                logging.exception("Fetch error")
                continue

            for article in articles['PubmedArticle']:
                try:
                    pmid = str(article['MedlineCitation']['PMID'])
                    article_data = article['MedlineCitation']['Article']
                    title = article_data.get("ArticleTitle", "No Title")
                    journal = article_data["Journal"]["Title"]
                    volume = article_data["Journal"]["JournalIssue"].get("Volume", "N/A")
                    issue = article_data["Journal"]["JournalIssue"].get("Issue", "N/A")
                    year = article_data["Journal"]["JournalIssue"]["PubDate"].get("Year", "N/A")
                    pages = article_data.get("Pagination", {}).get("MedlinePgn", "N/A")

                    doi = extract_doi(article_data.get("ELocationID", []))

                    for author in article_data.get("AuthorList", []):
                        if "ForeName" in author and "LastName" in author:
                            full_name = f"{author['LastName']}, {author['ForeName']}"

                            for aff in author.get("AffiliationInfo", []):
                                aff_text = aff.get("Affiliation", "")
                                email = extract_email(aff_text)
                                if not email or any(email.lower().endswith(f"@{domain}") for domain in PERSONAL_EMAIL_DOMAINS):
                                    continue

                                matched_country = next(
                                    (country for country in selected_countries
                                     if re.search(rf'\b{re.escape(country)}\b', aff_text, re.IGNORECASE)),
                                    None
                                )
                                if not matched_country:
                                    continue

                                university = extract_university_name(aff_text)
                                if not university:
                                    continue

                                mla = format_mla([full_name], title, journal, volume, issue, year, pages, doi)

                                data.append({
                                    "PMID": pmid,
                                    "Author": full_name,
                                    "Email": email,
                                    "Country": matched_country,
                                    "University": university,
                                    "Affiliation": aff_text,
                                    "MLA Citation": mla
                                })
                                break
                except Exception as e:
                    logging.exception("Article processing error")
                    st.warning(f"⚠️ Skipped an article due to error: {e}")
                    continue

            sleep(0.5)
            progress.progress((i + 1) / ((len(pmids) - 1) // BATCH_SIZE + 1))

        if data:
            df = pd.DataFrame(data)
            st.success(f"✅ Completed! {len(data)} valid entries extracted.")
            st.dataframe(df)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name="Results")
            output.seek(0)

            st.download_button(
                label="📁 Download Excel",
                data=output,
                file_name="pubmed_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("❌ No valid data extracted. Try refining your search.")

# =====================
# 📰 Google News Search
# =====================
elif menu == "Google News Search":
    st.title("📰 Google News Article Finder")

    query = st.text_input("Enter a topic, company, or keyword", "BYD Auto")
    max_articles = st.slider("Number of articles to display", 1, 20, 5)

    if st.button("Search News"):
        with st.spinner("Fetching news..."):
            news_results = get_google_news(query, max_articles=max_articles)

        if news_results:
            st.success(f"✅ Found {len(news_results)} articles.")
            for i, article in enumerate(news_results, 1):
                st.markdown(f"**{i}. [{article['title']}]({article['link']})**")
                st.markdown(f"*Published:* {article['published']}\n")
        else:
            st.warning("⚠️ No news articles found or feed could not be loaded.")

# =====================
# 📂 Excel Splitter
# =====================
elif menu == "Excel Splitter":
    st.title("📂 Excel Splitter - Split Large Excel File into Smaller Files")

    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

    entries_per_file = st.number_input(
        "Number of entries per file",
        min_value=100, max_value=10000, value=1500, step=100
    )

    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        total_rows = len(df)
        st.info(f"✅ Uploaded file has **{total_rows}** rows.")

        if st.button("🔨 Split Excel File"):
            with st.spinner("Splitting file..."):
                zip_buffer = BytesIO()
                with ZipFile(zip_buffer, "w") as zip_file:
                    num_files = (len(df) // entries_per_file) + (1 if len(df) % entries_per_file != 0 else 0)

                    for i in range(num_files):
                        start_row = i * entries_per_file
                        end_row = (i + 1) * entries_per_file
                        df_subset = df[start_row:end_row]

                        file_buffer = BytesIO()
                        df_subset.to_excel(file_buffer, index=False)
                        file_buffer.seek(0)

                        zip_file.writestr(f'split_part_{i + 1}.xlsx', file_buffer.read())

                zip_buffer.seek(0)

                st.download_button(
                    label="📦 Download ZIP of Split Files",
                    data=zip_buffer,
                    file_name="split_excel_files.zip",
                    mime="application/zip"
                )
# ===============================
# 🔎 Yahoo Finance Company Lookup
# ===============================
elif menu == "Yahoo Finance Company Lookup":
    st.title("🔎 Yahoo Finance Company Lookup")

    try:
        from yahooquery import search, Ticker
    except ImportError:
        st.error("❌ yahooquery is not installed. Please install it using `pip install yahooquery`")
        st.stop()

    def search_company_yahoo(name, limit=3):
        try:
            results = search(name)
            if not results or "quotes" not in results:
                return []
            companies = []
            for item in results['quotes']:
                if 'symbol' in item and 'shortname' in item:
                    companies.append({
                        "symbol": item['symbol'],
                        "name": item.get('shortname'),
                        "exchange": item.get('exchDisp'),
                        "type": item.get('typeDisp')
                    })
                    if len(companies) >= limit:
                        break
            return companies
        except Exception as e:
            return {"error": str(e)}

    def get_company_info(symbol):
        try:
            ticker = Ticker(symbol)
            profile = ticker.asset_profile
            info = profile.get(symbol, {})
            return info
        except Exception as e:
            return {"error": str(e)}

    st.markdown("Search for a company and fetch detailed metadata from Yahoo Finance.")
    company_input = st.text_input("Enter Company Name", value="Apple")

    if st.button("Search Company"):
        if not company_input.strip():
            st.warning("Please enter a company name.")
            st.stop()

        with st.spinner("Searching Yahoo Finance..."):
            matches = search_company_yahoo(company_input)

        if isinstance(matches, dict) and "error" in matches:
            st.error(f"Error: {matches['error']}")
        elif not matches:
            st.warning("No matches found.")
        else:
            st.success(f"Found {len(matches)} match(es). Showing top result.")

            top_match = matches[0]
            st.markdown(f"**Top Match:** `{top_match['name']}`")
            st.markdown(f"- **Symbol:** {top_match['symbol']}")
            st.markdown(f"- **Exchange:** {top_match['exchange']}")
            st.markdown(f"- **Type:** {top_match['type']}")

            with st.spinner("Fetching company details..."):
                details = get_company_info(top_match['symbol'])

            if isinstance(details, dict) and "error" in details:
                st.error(f"Error: {details['error']}")
            elif details:
                st.subheader("📊 Company Details")
                for key, value in details.items():
                    st.markdown(f"- **{key}:** {value}")
            else:
                st.warning("No detailed data available for this symbol.")
# ===============================
# 🧾 Excel Merger-Flatten Viewer
# ===============================
elif menu == "Excel Merger-Flatten Viewer":
    st.title("🧾 Flatten Excel with Merged Cells")

    uploaded_file = st.file_uploader("Upload an Excel file with merged cells", type=["xlsx"])

    if uploaded_file:
        from openpyxl import load_workbook

        with st.spinner("Processing..."):
            wb = load_workbook(uploaded_file)
            ws = wb.active

            # Step 1: Read values into 2D list
            data = [[cell.value for cell in row] for row in ws.iter_rows()]

            # Step 2: Fill in merged cell ranges
            for merged_range in ws.merged_cells.ranges:
                min_row, min_col, max_row, max_col = (
                    merged_range.min_row,
                    merged_range.min_col,
                    merged_range.max_row,
                    merged_range.max_col
                )
                top_left_value = ws.cell(row=min_row, column=min_col).value

                for row in range(min_row - 1, max_row):
                    for col in range(min_col - 1, max_col):
                        if row != min_row - 1 or col != min_col - 1:
                            data[row][col] = top_left_value

            # Step 3: Convert to DataFrame and display
            df = pd.DataFrame(data)
            st.success("✅ File processed. Here's a preview:")
            st.dataframe(df)

            # Step 4: Download cleaned Excel
            output_buffer = BytesIO()
            df.to_excel(output_buffer, index=False, header=False)
            output_buffer.seek(0)

            st.download_button(
                label="📥 Download Flattened Excel",
                data=output_buffer,
                file_name="flattened_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# ===============================
# 🔍 Fuzzy Name Matcher
# ===============================
elif menu == "Fuzzy Name Matcher":
    st.title("🔍 Fuzzy Name Matcher")
    st.markdown("Match standardized names against raw names using fuzzy logic.")

    # === 1. Downloadable Excel Template ===
    def generate_standard_template():
        from io import BytesIO
        df_template = pd.DataFrame({
            "Raw Data": ["Tesla Motors Inc.", "Massachusetts General Hospital", "Harvard Univ.", "Amazon.com LLC"],
            "Standardized Names": ["Tesla", "Massachusetts General", "Harvard University", "Amazon"]
        })
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            df_template.to_excel(writer, index=False, sheet_name="Template")
        buffer.seek(0)
        return buffer

    st.markdown("📄 **Download the standard Excel template** and use it to upload your data.")
    st.download_button(
        label="📥 Download Standard Template",
        data=generate_standard_template(),
        file_name="fuzzy_match_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # === 2. User Inputs ===
    uploaded_file = st.file_uploader("📤 Upload Excel File", type=["xlsx"])

    threshold = st.slider("🧪 Match Threshold", min_value=70, max_value=100, value=90, step=1)

    user_keywords = st.text_input("🚫 Enter keywords to ignore (comma-separated)", value="inc, ltd, university, hospital")

    reverse_match = st.checkbox("🔄 Reverse Matching Direction (Raw ⇄ Standardized)", value=False)

    # === 3. Processing ===
    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        df.columns = df.columns.str.strip()
        st.success(f"✅ Uploaded file with {df.shape[0]} rows and {df.shape[1]} columns.")

        st.dataframe(df.head())

        raw_col = st.selectbox("📌 Select Raw Column", df.columns)
        standard_col = st.selectbox("🎯 Select Standardized Column", df.columns)

        if st.button("🚀 Run Fuzzy Matching"):
            with st.spinner("Matching in progress..."):

                from rapidfuzz import fuzz
                import re
                from io import BytesIO

                # === 4. Prepare Stopwords ===
                default_stopwords = {
                    'inc', 'inc.', 'ltd', 'ltd.', 'co', 'co.', 'corporation', 'corp', 'corp.', 'llc', 'plc', 'limited',
                    'university', 'college', 'hospital', 'center', 'institute', 'school', 'company', 'system',
                    'technology', 'medical', 'science', 'sciences', 'people', 'peoples'
                }
                user_stopwords = {w.strip().lower() for w in user_keywords.split(",") if w.strip()}
                STOPWORDS = default_stopwords.union(user_stopwords)

                def clean_name(name):
                    name = name.lower()
                    name = re.sub(r'[^\w\s]', '', name)
                    words = name.split()
                    return ' '.join([w for w in words if w not in STOPWORDS])

                # === 5. Flip direction if needed ===
                if reverse_match:
                    raw_col, standard_col = standard_col, raw_col

                raw_original = df[raw_col].dropna().astype(str).unique().tolist()
                standard_original = df[standard_col].dropna().astype(str).unique().tolist()

                raw_cleaned = [clean_name(name) for name in raw_original]
                standard_cleaned = [clean_name(name) for name in standard_original]

                matched_standardized, matched_raw, match_strengths = [], [], []
                borderline_matches, unmatched_standardized = [], []
                matched_raw_indices = set()

                for std_orig, std_clean in zip(standard_original, standard_cleaned):
                    best_score = 0
                    best_index = None

                    for idx, raw_clean in enumerate(raw_cleaned):
                        score1 = fuzz.token_sort_ratio(std_clean, raw_clean)
                        score2 = fuzz.token_set_ratio(std_clean, raw_clean)
                        avg_score = (score1 + score2) / 2

                        if (score1 >= threshold and score2 >= threshold) or avg_score >= (threshold + 5):
                            if avg_score > best_score:
                                best_score = avg_score
                                best_index = idx

                    if best_index is not None:
                        raw_orig = raw_original[best_index]
                        matched_standardized.append(std_orig)
                        matched_raw.append(raw_orig)
                        matched_raw_indices.add(best_index)

                        if best_score >= 100:
                            match_strengths.append("Exact Match")
                        elif best_score >= threshold:
                            match_strengths.append(f"{round(best_score, 1)}% Match")
                        elif best_score >= (threshold - 5):
                            borderline_matches.append((std_orig, raw_orig, f"{round(best_score, 1)}% Match"))
                    else:
                        unmatched_standardized.append(std_orig)

                unmatched_raw = [raw_original[i] for i in range(len(raw_original)) if i not in matched_raw_indices]

                # === 6. Export to Excel ===
                output_buffer = BytesIO()
                with pd.ExcelWriter(output_buffer) as writer:
                    pd.DataFrame({
                        'Matched Standardized Name': matched_standardized,
                        'Matched Raw Name': matched_raw,
                        'Match Strength': match_strengths
                    }).to_excel(writer, sheet_name='Matches', index=False)

                    pd.DataFrame({'Unmatched Standardized Names': unmatched_standardized}).to_excel(
                        writer, sheet_name='Unmatched Standardized', index=False
                    )
                    pd.DataFrame({'Unmatched Raw Names': unmatched_raw}).to_excel(
                        writer, sheet_name='Unmatched Raw', index=False
                    )
                    if borderline_matches:
                        pd.DataFrame(borderline_matches, columns=['Standardized Name', 'Potential Raw Match', 'Match Strength']).to_excel(
                            writer, sheet_name='Borderline Matches', index=False
                        )

                output_buffer.seek(0)
                st.success("✅ Matching completed!")
                st.download_button(
                    label="📥 Download Results Excel",
                    data=output_buffer,
                    file_name="fuzzy_match_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# ===============================
# 🗂️ Categories Lister
# ===============================
elif menu == "Categories Lister":
    st.title("🗂️ Categories Lister")
    st.markdown(
        "Extract **all category values** from an Excel file (including merged cells) "
        "and export them as a **unique category list**."
    )

    uploaded_file = st.file_uploader(
        "📤 Upload Excel File (with merged category columns)",
        type=["xlsx"]
    )

    if uploaded_file:
        try:
            df = pd.read_excel(uploaded_file)
            st.success(f"✅ File loaded successfully ({df.shape[0]} rows × {df.shape[1]} columns)")
            st.dataframe(df.head())

            if st.button("🚀 Extract Categories"):
                with st.spinner("Processing categories..."):

                    # Forward-fill merged cells COLUMN-WISE
                    df_filled = df.ffill()

                    # Extract all values from all columns
                    all_categories = (
                        df_filled.astype(str)
                                 .values
                                 .flatten()
                    )

                    # Clean: remove blanks, nan, duplicates
                    categories = (
                        pd.Series(all_categories)
                          .replace(["nan", "None", ""], pd.NA)
                          .dropna()
                          .unique()
                    )

                    result_df = pd.DataFrame({"Category": categories})

                    st.success(f"✅ Extracted {len(result_df)} unique categories")
                    st.dataframe(result_df)

                    # Download Excel
                    output_buffer = BytesIO()
                    with pd.ExcelWriter(output_buffer, engine="xlsxwriter") as writer:
                        result_df.to_excel(writer, index=False, sheet_name="Categories")

                    output_buffer.seek(0)

                    st.download_button(
                        label="📥 Download Categories Excel",
                        data=output_buffer,
                        file_name="categories_list.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

        except Exception as e:
            st.error(f"❌ Failed to process file: {e}")

# ===============================
# 📄 Patent Scraper
# ===============================
elif menu == "Patent Scraper":
    st.title("📄 Patent Scraper (Google Patents)")

    import requests
    from bs4 import BeautifulSoup
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from deep_translator import GoogleTranslator
    from threading import Lock

    base_url = "https://patents.google.com/patent/{}"
    MAX_WORKERS = 6

    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})

    translation_cache = {}
    cache_lock = Lock()

    # --------------------------
    # Helpers
    # --------------------------
    def is_english(text):
        try:
            text.encode('ascii')
            return True
        except:
            return False

    def translate_batch(text_list):
        results = []
        to_translate = []
        index_map = {}

        for i, text in enumerate(text_list):
            text = str(text).strip()

            if not text:
                results.append("")
                continue

            with cache_lock:
                if text in translation_cache:
                    results.append(translation_cache[text])
                    continue

            if is_english(text):
                with cache_lock:
                    translation_cache[text] = text
                results.append(text)
            else:
                index_map[len(to_translate)] = i
                to_translate.append(text)
                results.append(None)

        if to_translate:
            try:
                translated = GoogleTranslator(source='auto', target='en').translate_batch(to_translate)
            except:
                translated = to_translate

            for idx, t in enumerate(translated):
                original_index = index_map[idx]
                original_text = text_list[original_index]

                with cache_lock:
                    translation_cache[original_text] = t

                results[original_index] = t

        return results

    def extract_pdf_link(soup):
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if href.endswith(".pdf") and "googleapis" in href:
                return href
        return ""

    def generate_alternate_doc_numbers(doc):
        alternatives = []

        if re.match(r"US\d{10,}A\d", doc):
            alternatives.append(doc[:6] + "0" + doc[6:])

        if doc.startswith("USD") and doc.endswith("S"):
            alternatives.append(doc + "1")
            alternatives.append(doc + "2")

        return alternatives

    def fetch_patent(doc):
        try:
            r = session.get(base_url.format(doc), timeout=20)
            if r.status_code == 200 and "Error 404" not in r.text:
                return r
        except:
            pass
        return None

    def process_patent(doc):
        inventor = ""
        assignee = ""
        status = ""
        adjusted_expiration = ""
        availability = "Available"
        final_doc_used = doc

        response = fetch_patent(doc)

        if not response:
            for alt in generate_alternate_doc_numbers(doc):
                response = fetch_patent(alt)
                if response:
                    final_doc_used = alt
                    break

        if not response:
            return {"document number": doc, "availability": "Not Found"}

        soup = BeautifulSoup(response.text, "html.parser")

        # INVENTOR
        inventors = []
        inventor_dt = soup.find("dt", string=re.compile("Inventor", re.I))
        if inventor_dt:
            node = inventor_dt.find_next_sibling()
            while node and node.name == "dd":
                for a in node.find_all("a"):
                    if a.text.strip():
                        inventors.append(a.text.strip())
                node = node.find_next_sibling()

        inventor = " | ".join(dict.fromkeys(inventors))
        inventor_translated = " | ".join(dict.fromkeys(translate_batch(inventors)))

        # ASSIGNEE
        assignees = []
        assignee_dt = soup.find("dt", string=re.compile("Assignee", re.I))
        if assignee_dt:
            node = assignee_dt.find_next_sibling()
            while node and node.name == "dd":
                for a in node.find_all("a"):
                    if a.text.strip():
                        assignees.append(a.text.strip())
                node = node.find_next_sibling()

        assignee = " | ".join(dict.fromkeys(assignees))
        assignee_translated = " | ".join(dict.fromkeys(translate_batch(assignees)))

        # STATUS
        status_tag = soup.find(string=re.compile("Status", re.I))
        if status_tag:
            try:
                raw_status = status_tag.find_next().text.strip()
                if "," in raw_status:
                    status = raw_status.split(",")[0].strip()
                    date_match = re.search(r"\d{4}-\d{2}-\d{2}", raw_status)
                    if date_match:
                        adjusted_expiration = date_match.group()
                else:
                    status = raw_status
            except:
                pass

        pdf = extract_pdf_link(soup)

        return {
            "document number": doc,
            "used document": final_doc_used,
            "inventor": inventor,
            "inventor (translated)": inventor_translated,
            "Current assignee": assignee,
            "assignee (translated)": assignee_translated,
            "status": status,
            "adjusted expiration": adjusted_expiration,
            "pdf": pdf,
            "availability": availability
        }

    # --------------------------
    # UI
    # --------------------------
    input_method = st.radio("Input Method", ["Paste", "Upload Excel"])

    patents = []

    if input_method == "Paste":
        text = st.text_area("Paste patents (comma or newline separated)")
        if text:
            patents = re.split(r"[,\n]+", text)
            patents = [p.strip() for p in patents if p.strip()]

    else:
        file = st.file_uploader("Upload Excel", type=["xlsx"])
        if file:
            df = pd.read_excel(file)
            col = st.selectbox("Select Patent Column", df.columns)
            patents = df[col].dropna().astype(str).tolist()

    columns = [
        "inventor",
        "inventor (translated)",
        "Current assignee",
        "assignee (translated)",
        "status",
        "adjusted expiration",
        "pdf"
    ]

    selected_cols = st.multiselect("Select Output Columns", columns, default=columns)

    if st.button("🚀 Run Patent Scraper") and patents:
        progress = st.progress(0)
        results = []

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(process_patent, doc): doc for doc in patents}

            for i, future in enumerate(as_completed(futures)):
                results.append(future.result())
                progress.progress((i + 1) / len(futures))

        df_out = pd.DataFrame(results)

        base_cols = ["document number", "used document", "availability"]
        final_cols = base_cols + selected_cols
        final_cols = [c for c in final_cols if c in df_out.columns]

        df_out = df_out[final_cols]

        st.success("✅ Scraping Completed")
        st.dataframe(df_out)

        buffer = BytesIO()
        df_out.to_excel(buffer, index=False)
        buffer.seek(0)

        st.download_button(
            "📥 Download Excel",
            buffer,
            file_name="patent_output.xlsx"
        )
